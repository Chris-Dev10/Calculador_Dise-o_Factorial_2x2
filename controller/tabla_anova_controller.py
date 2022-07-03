from view.tabla_anova_view import TablaAnovaView
from model.tabla_anova_model import TablaAnovaModel

import numpy
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from scipy.stats import f


class TablaAnovaController(object):
    def __init__(self):
        self._tabla_anova_v = TablaAnovaView()
        self._tabla_anova_m = TablaAnovaModel(self.load_datos_from_excel())
        self._suma_cuadrados = 0; self._grados_libertad = 0; self._cuadrados_medios = 0; self._fs_calculadas = 0; self._fs_criticas = 0
        self._tabla_anova_widget = None
    

    @property
    def get_tabla_anova_v(self):
        return self._tabla_anova_v


    def load_datos_from_excel(self):
        datos = numpy.array([[0, 0, 0], [0, 0, 0], [0, 0, 0], [0, 0, 0], [0, 0, 0], [0, 0, 0], [0, 0, 0], [0, 0, 0], [0, 0, 0], [0, 0, 0], [0, 0, 0], [0, 0, 0]], numpy.int16)
        
        datos_workbook = load_workbook('model/Datos.xlsx')
        ejemplo_1_worksheet = datos_workbook['Ejemplo 1']

        for row in range(0, 12):
            for column in range(0, 3):
                column_number = column + 1
                column_letter = get_column_letter(column_number)

                row_number = row + 1
                datos[row][column] = ejemplo_1_worksheet[column_letter + str(row_number)].value
        
        return datos


    def calculate_datos_tabla_anova(self):
        SCA = self._tabla_anova_m.calculate_suma_cuadrados_de_factor('A')
        SCB = self._tabla_anova_m.calculate_suma_cuadrados_de_factor('B')
        SCAB = self._tabla_anova_m.calculate_suma_cuadrados_interaccion_entre_factores(SCA, SCB)
        SCT = self._tabla_anova_m.calculate_suma_cuadrados_totales()
        SCE = self._tabla_anova_m.calculate_suma_cuadrados_del_error(SCT, SCA, SCB, SCAB)

        self._suma_cuadrados = (round(SCA), round(SCB), round(SCAB), round(SCE), round(SCT))

        GLA = self._tabla_anova_m.get_niveles_factor_a - 1
        GLB = self._tabla_anova_m.get_niveles_factor_b - 1
        GLAB = GLA * GLB
        GLE = self._tabla_anova_m.get_niveles_factor_a * self._tabla_anova_m.get_niveles_factor_b * (self._tabla_anova_m.get_replicas_por_nivel - 1)
        GLT = self._tabla_anova_m.get_replicas_totales - 1

        self._grados_libertad = (GLA, GLB, GLAB, round(GLE), GLT)

        CMA = SCA / GLA
        CMB = SCB / GLB
        CMAB = SCAB / GLAB
        CME = SCE / GLE

        self._cuadrados_medios = (round(CMA, 2), round(CMB, 2), round(CMAB, 2), round(CME, 2))

        FCALA = CMA / CME
        FCALB = CMB / CME
        FCALAB = CMAB / CME

        self._fs_calculadas = (round(FCALA, 2), round(FCALB, 2), round(FCALAB, 2))

        FCRITA = f.ppf(self._tabla_anova_m.get_alfa, GLA, GLE)
        FCRITB = f.ppf(self._tabla_anova_m.get_alfa, GLB, GLE)
        FCRITAB = f.ppf(self._tabla_anova_m.get_alfa, GLAB, GLE)

        self._fs_criticas = (round(FCRITA, 2), round(FCRITB, 2), round(FCRITAB, 2))
    

    def create_tabla_anova_widget(self):
        self._tabla_anova_widget = self._tabla_anova_v.create_tabla_anova_widget(self._suma_cuadrados, self._grados_libertad, self._cuadrados_medios, self._fs_calculadas, self._fs_criticas)
    

    def add_tabla_anova_widget_to_view(self):
        self._tabla_anova_v.add_widget(self._tabla_anova_widget)